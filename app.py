import asyncio
import os
import re
import tempfile
import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
import datetime
import time
import zipfile
import tarfile
from selenium.common.exceptions import TimeoutException 
from io import BytesIO
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import StaleElementReferenceException 
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

### SELENIUM FUNCTIONS
def set_date(key1, key2):
            date_start, date_end = st.columns(2)
            with date_start:
                start_date = st.date_input(
                    "Ngày bắt đầu:", 
                    format="DD/MM/YYYY", 
                    key= key1).strftime("%d/%m/%Y")
            with date_end:
                end_date = st.date_input(
                    "Ngày kết thúc:",  
                    format="DD/MM/YYYY",
                    key= key2).strftime("%d/%m/%Y")
            return start_date, end_date

def enter_dates(wait, start_date, end_date, btn_path):
    """
    Enters the start and end date in the appropriate fields.
    """
    date_btn = wait.until(EC.presence_of_all_elements_located((By.XPATH, btn_path)))
    # date_btn = driver.find_elements(By.XPATH, btn_path)
    date_btn[0].clear()
    date_btn[0].send_keys(start_date)

    date_btn[1].clear()
    date_btn[1].send_keys(end_date)

def selenium_web_driver(temp_folder):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": temp_folder,  # Ensure this path matches your temp dir
        "download.prompt_for_download": False, 
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })
    # Point the browser to the correct location
    chrome_options.binary_location = "/usr/bin/chromium"

    # Use chromedriver installed by the system package manager
    driver = webdriver.Chrome(service=Service("/usr/bin/chromedriver"), options=chrome_options)
    action=ActionChains(driver,10)
    wait = WebDriverWait(driver, 20)
    
    return driver, action, wait

def wait_for_download(temp_folder, timeout=30):
    '''Wait for a file to be downloaded to the download path'''
    start_time = time.time()
    while time.time() - start_time < timeout:
        files = os.listdir(temp_folder)
        if files:
            # Get the most recently downloaded file
            latest_file = max([os.path.join(temp_folder, f) for f in files], key=os.path.getctime)
            
            # Ensure the file is fully downloaded (not a .crdownload file)
            if not latest_file.endswith('.crdownload') and os.path.getsize(latest_file) > 0:
                return latest_file
        time.sleep(2)  # Slightly increase sleep to reduce CPU load

    return None

def stream_data(data):
    for word in data.split(" "):
        yield word + " "
        time.sleep(0.03)

def download_tar(temp_folder):
                # Define the tar archive path
                tar_path = os.path.join(temp_folder, 'XML_files.tar')
                
                # Create the tar archive in the download path
                with tarfile.open(tar_path, 'w') as tar:
                    for file in os.listdir(temp_folder):
                        tar.add(os.path.join(temp_folder, file), arcname=file)

                # Provide download of the tar file
                with open(tar_path, 'rb') as f:
                    if st.download_button("Tải thư mục XML", f, file_name="XML_files.tar", type="primary"):
                        downloading_message = 'Đang tải thư mục ...'
                        progress_bar = st.progress(0, text=downloading_message)
                        for percent_complete in range(100):
                            time.sleep(0.01)
                            progress_bar.progress(percent_complete + 1, text=downloading_message)
                        time.sleep(1)
                        st.success("Đã tải thư mục XML thành công")

### TAB 3 FUNCTIONS
def convert_date_format(date_str):
    # Parse the date string
    date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')

    # Format the date in the desired format
    formatted_date = f"Ngày {date_obj.day:02} tháng {date_obj.month:02} năm {date_obj.year}"
    return formatted_date

def show_progress_bar(message):
    """Function to display a progress bar and a success message."""
    progress_bar = st.progress(0, text=message)
    for percent_complete in range(100):
        time.sleep(0.01)
        progress_bar.progress(percent_complete + 1, text=message)
    time.sleep(1)
    return progress_bar

def download_success_handler(success_key, message):
    """Handles download success for PXK and PTT."""
    if st.session_state[success_key]:
        downloading_message = message
        progress_bar = show_progress_bar(downloading_message)
        st.success(f"{message.split(' ')[1]} thành công")
        progress_bar.empty()
        st.session_state[success_key] = False

def create_download_button(label, buffer, file_name, key):
    """Creates a download button with specified parameters."""
    st.download_button(
        on_click=download,
        type="primary",
        label=label,
        data=buffer,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key
    )

@st.cache_data
def pxk_data_from_xml(file):
    data=[]
    ggia = 0
    tree = ET.parse(file)
    root = tree.getroot()
    shdon = int(root.find('.//TTChung/SHDon').text) 
    nmua = root.find('.//NMua/Ten').text 
    nmua_dc = root.find('.//NMua/DChi').text
    nban = (root.find('.//NBan/Ten').text).replace('HỘ KINH DOANH', '').strip().title()
    nban_mst = root.find('.//NBan/MST').text
    nban_dc = (root.find('.//NBan/DChi').text).replace(', Bà Rịa - Vũng Tàu', '').strip()
    date = convert_date_format(root.find('.//NLap').text) 
    tbc = root.find('.//TgTTTBChu').text 
    ts = int(root.find('.//TgTTTBSo').text)
    for item in root.findall('.//HHDVu'):
        stt = int(item.find('STT').text) if item.find('STT') is not None else ""
        thhdv = item.find('THHDVu').text if item.find('THHDVu') is not None else ""
        dvtinh = item.find('DVTinh').text if item.find('DVTinh') is not None else ""
        sluong = float(item.find('SLuong').text.replace(',', '')) if item.find('SLuong') is not None and item.find('SLuong').text else 0
        dgia = int(item.find('DGia').text.replace(',', '').replace('.', '')) if item.find('DGia') is not None and item.find('DGia').text else 0
        thtien = int(item.find('ThTien').text.replace(',', '').replace('.', '')) if item.find('ThTien') is not None and item.find('ThTien').text else 0
        if "Đã giảm" in thhdv:
            ggia_number = re.search(r'(\d{1,3}(?:\.\d{3})*)\s+đồng', thhdv)
            ggia = int(ggia_number.group(1).replace('.', ''))
            data.append([stt, thhdv, dvtinh, sluong, dgia, thtien, shdon, ggia])
        else:
            data.append([stt, thhdv, dvtinh, sluong, dgia, thtien, shdon])

    return  shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts, ggia, data

def create_pxk(all_data):
                        with st.spinner("Đang tạo phiếu ..."):  # Display loading spinner during process
                            time.sleep(3)  # Simulate processing delay

                            pxk_file_path = 'pxk.xlsx'  # Path to PXK Excel file
                            ptt_file_path = 'ptt.xlsx'  # Path to PTT Excel file

                            pxk_wb = load_workbook(pxk_file_path)  # Load PXK Excel workbook
                            ptt_wb = load_workbook(ptt_file_path)  # Load PTT Excel workbook

                            # Populate PXK workbook with data from all_data
                            for shdon, nmua, _, nban, nban_dc, nban_mst, date, tbc, ts, ggia, df in all_data:
                                pxk_excel(pxk_wb, shdon, nmua, nban, nban_dc, nban_mst, date, tbc, ggia, df)

                            # Populate PTT workbook with data from all_data
                            for shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts, ggia, df in all_data:    
                                ptt_excel(ptt_wb, shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts)

                            # Set the first sheet as active if there are more than one sheet, and remove template sheet
                            if len(pxk_wb.sheetnames) > 1 and len(ptt_wb.sheetnames) > 1:
                                pxk_wb.active = 1  # Set second sheet as active in PXK workbook
                                ptt_wb.active = 1  # Set second sheet as active in PTT workbook
                                pxk_wb.remove(pxk_wb['Template'])  # Remove 'Template' sheet from PXK workbook
                                ptt_wb.remove(ptt_wb['Template'])  # Remove 'Template' sheet from PTT workbook

                            # Save PXK workbook to a buffer in memory
                            pxk_buffer = BytesIO()
                            pxk_wb.save(pxk_buffer)
                            pxk_buffer.seek(0)  # Move buffer cursor to the beginning

                            # Save PTT workbook to a buffer in memory
                            ptt_buffer = BytesIO()
                            ptt_wb.save(ptt_buffer)
                            ptt_buffer.seek(0)  # Move buffer cursor to the beginning
                        return pxk_buffer, ptt_buffer

def display_pxk(shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts, ggia, data):
    columns = ['STT', 'Tên hàng hóa, dịch vụ', 'Đơn vị tính', 'Số lượng', 'Đơn giá', 'Thành tiền', 'Số hóa đơn']
    if ggia:
        columns.append('Giảm giá')

    df = pd.DataFrame(data, columns=columns)
    df = df[df['Thành tiền'] != 0]
    df.loc[:, 'Tên hàng hóa, dịch vụ'] = df['Tên hàng hóa, dịch vụ'].str.capitalize()
    df.loc[:, 'Đơn vị tính'] = df['Đơn vị tính'].str.capitalize()

    df = df[[
        'STT', 
        'Tên hàng hóa, dịch vụ', 
        'Đơn vị tính', 
        'Số lượng', 
        'Đơn giá', 
        'Thành tiền'
    ]]

    # all_data.append((shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts, ggia, df))

    st.subheader(f"Số hóa đơn: {shdon}")
    st.text(f"Ngày-Tháng-Năm: {date}")
    st.text(f"Tên khách: {nmua}")


    st.dataframe(
        df.style.format({
        'Số lượng': lambda x: f'{x:.1f}'.replace('.', ','),
        'Đơn giá': lambda x: f"{x:,.0f}".replace(',', '.'),
        'Thành tiền': lambda x: f"{x:,.0f}".replace(',', '.')}),
        width=800
    )
    
    st.text(f"Giảm giá: {ggia} đồng")
    st.text(f"Tổng thành tiền: {float(ts)}")
    st.text(f"Tổng thành tiền chữ: {tbc}")

    return df

def pxk_excel(wb, shdon, nmua, nban, nban_dc, nban_mst, date, tbc, ggia, df):
    template_sheet = wb['Template']
    
    # Create a new sheet by copying the template sheet
    new_sheet_name = f"{shdon}"  # Ensure the sheet name is within 31 characters
    wb.copy_worksheet(template_sheet).title = new_sheet_name
    ws = wb[new_sheet_name]

    if 'Địa chỉ' in template_sheet['A2'].value:
        parts = template_sheet['A2'].value.split('\n', 1) 
        new_value = f"{parts[0].strip()} {nban_dc}\n{parts[1].strip()} {nban_mst}"
    ws['A2'] = new_value
    ws['A1'] = f"Hộ kinh doanh: {nban}"
    ws['A4'] = date
    ws['A8'] = f"Địa điểm xuất kho: {nban_dc}"
    ws['C41'] = shdon

    if nban == "An Vinh":
        ws['C45'] = "Phan Duy Thanh"
        ws['F45'] = "Phan Duy Thanh"
    elif nban == "Tuyết Oanh":
        ws['C45'] = "Thanh Thúy"
        ws['F45'] = "Thanh Thúy"
    else:
        ws['C45'] = nban
        ws['F45'] = nban
    ws['C40'] = tbc
    ws['A6'] = f"Tên đơn vị mua hàng: {nmua}"

    if ggia:
        ws['D39'] = ggia 

    start_row = 11
    start_column = 1  # Column A
    for i, row_data in enumerate(df.itertuples(index=False), start=start_row):
        if row_data[2]:
            ws.cell(row=i, column=start_column + 0, value=row_data[0])  # STT
            ws.cell(row=i, column=start_column + 1, value=row_data[1])  # Tên hàng hóa, dịch vụ
            ws.cell(row=i, column=start_column + 2, value=row_data[2])  # Đơn vị tính
            ws.cell(row=i, column=start_column + 3, value=row_data[3])  # Số lượng
            ws.cell(row=i, column=start_column + 4, value=row_data[4])  # Đơn giá
            ws.cell(row=i, column=start_column + 5, value=row_data[5])  # Thành tiền
def ptt_excel(wb, shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts):

# def display_ptt(shdon, nmua, date, tbc, ts):
#     st.subheader(f"Số hóa đơn: {shdon}")
#     st.text(f"Ngày-Tháng-Năm: {date}")
#     st.text(f"Tên khách: {nmua}")
#     st.text(f"Tổng tiền: {ts} đồng")
#     st.text(f"Tổng iền bằng chữ: {tbc}")

    template_sheet = wb['Template']
    new_sheet_name = f"{shdon}"  
    wb.copy_worksheet(template_sheet).title = new_sheet_name
    ws = wb[new_sheet_name]

    if 'Địa chỉ' in template_sheet['A2'].value:
        parts = template_sheet['A2'].value.split('\n', 1) 
        new_value = f"{parts[0].strip()} {nban_dc}\n{parts[1].strip()} {nban_mst}"
    ws['A7'] = f"Địa chỉ: {nmua_dc}"
    ws['A2'] = new_value
    ws['A1'] = f"Hộ kinh doanh: {nban}"
    if nban == "An Vinh":
        ws['A17'] = "Duy Thanh"
        ws['C17'] = "Duy Thanh"
    elif nban == "Tuyết Oanh":
        ws['A17'] = "Thanh Thúy"
        ws['C17'] = "Thanh Thúy"
    else:
        ws['A17'] = nban
        ws['C17'] = nban
    ws['B4'] = date
    ws['F14'] = date
    ws['C10'] = tbc
    ws['C24'] = tbc
    ws['B9'] = ts
    ws['C23'] = ts
    ws['D6'] = nmua
    ws['D12'] = shdon

### TAB 1 FUNCTIONS
def extract_number_vnpt(string):
    # Find the position of the last underscore and the '.zip' extension
    last_underscore_pos = string.rfind('_')
    dot_zip_pos = string.find('.zip')
    
    # Extract the number between the last underscore and '.zip'
    number = string[last_underscore_pos + 1:dot_zip_pos]
    return number

def extract_zipfile(zip_file, extract_to):
    extracted_files = []
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        for file in zip_ref.namelist():
            if file.endswith(".xml"):
                extracted_files.append(file)
                zip_ref.extract(file, extract_to)
                
    return extracted_files

def download_icon_vnpt(driver, action, wait, temp_folder):
    placeholder=st.empty()
    xml_files = []
    try:
        icons = wait.until(
            EC.presence_of_all_elements_located((By.XPATH, "//a[@title='Xem chi tiết hóa đơn']"))
        )
        icons = icons[:len(icons)//2]  
        time.sleep(3)

        seen_files = set()
        for icon in icons:
            # Move to the icon and click to open details
            action.move_to_element(icon).perform()
            driver.execute_script("arguments[0].click();", icon)
            time.sleep(3)

            download_button = driver.find_element(By.XPATH, "//div[@id='taiXml']")
            # download_button = wait.until(
            #     EC.presence_of_element_located((By.XPATH, "//div[@id='taiXml']")))

            if download_button:
                driver.execute_script("arguments[0].click();", download_button)
                time.sleep(3)
                downloaded_file = wait_for_download(temp_folder)

                if downloaded_file:
                    shd = extract_number_vnpt(os.path.basename(downloaded_file))
                    extracted_files = extract_zipfile(downloaded_file, temp_folder)

                    placeholder.write_stream(stream_data(f"Đang tải hóa đơn số {shd} ..."))

                    for file in extracted_files:
                        xml_file = shd + file[file.index('.xml'):]
                        file_path = os.path.join(temp_folder, file)

                        if xml_file in seen_files:
                            os.remove(file_path)
                        else:
                            seen_files.add(xml_file)
                            xml_files.append((xml_file, file))
                            os.rename(file_path, os.path.join(temp_folder, xml_file))
            

            close_button = wait.until(
                EC.presence_of_element_located((By.XPATH, "//button[@class='close']"))
            )
            driver.execute_script("arguments[0].click();", close_button)
            time.sleep(2)

        # Check if any XML files were downloaded
        if xml_files:
            return xml_files
        else:
            st.write_stream(stream_data("Không có hóa đơn nào được tải xuống"))
            return None
        
    except StaleElementReferenceException:
        return download_icon_vnpt(driver, action, wait, temp_folder)

    except TimeoutException:
        st.write_stream(stream_data(f"Không tìm thấy hóa đơn"))
        return None

def handle_vnpt_download(driver, action, wait, user, start_date, end_date, temp_folder):          
    with st.status("Đang tải XML tự động ...", expanded=True) as status:
        try:
            driver.get('https://hkd.vnpt.vn/account/login')

            wait.until(
                EC.presence_of_element_located((By.CLASS_NAME, 'form-horizontal'))
            )

            username = driver.find_element(By.NAME, 'UserName')
            password = driver.find_element(By.NAME, 'Password')

            if user == "Trần Minh Đạt":
                username.send_keys(os.getenv('username'))
                password.send_keys(os.getenv('password'))
                password.send_keys(Keys.RETURN)
            else:
                username.send_keys(os.getenv('username_2'))
                password.send_keys(os.getenv('password_2'))
                password.send_keys(Keys.RETURN)
                
            st.write_stream(stream_data((f"Đang đăng nhập tài khoản {user}...")))

            qlhd = wait.until(
                EC.presence_of_element_located((By.XPATH,"//a[@href='/Thue/QuanLyHoaDon']"))
            )
            driver.execute_script("arguments[0].click();", qlhd)
            st.write_stream(stream_data(("Đang vào mục Quản Lý Hóa Đơn ...")))

            enter_dates(wait, start_date, end_date, btn_path="//input[@class='dx-texteditor-input']")

            search_btn = wait.until(
                EC.presence_of_all_elements_located((By.CLASS_NAME, "dx-button-content"))
            )
            search_btn[3].click()
            time.sleep(3)
            st.write_stream(stream_data(("Đang tìm hóa đơn ...")))

            page_indexes = driver.find_element(By.XPATH, "//div[@class='dx-page-indexes']")
            all_pages = page_indexes.find_elements(By.CLASS_NAME, "dx-page")
            st.write_stream(stream_data((f"Tổng số trang: {len(all_pages)}")))

            # next_btn = page_indexes.find_element(By.XPATH, "//div[@aria-label='Next page']")
            
            final_xml_files = []
            page_index = 0

            while page_index < len(all_pages):
                try:
                    xml_files = download_icon_vnpt(driver, action, wait, temp_folder)
                    if xml_files:
                        final_xml_files.extend(xml_files)

                        # next_btn = page_indexes.find_element(By.XPATH, "//div[@aria-label='Next page']")
                        # st.write(next_btn)
                    next_button = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//div[@aria-label='Next page']"))
                    )
                    driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
                    driver.execute_script("arguments[0].click();", next_button)
                    page_index += 1
                except TimeoutException:
                    break  
                except Exception as e:
                    st.write(f"Lỗi tải: {e}")
                    return None
            

            for f in os.listdir(temp_folder):
                if f.endswith('.zip'):
                    os.remove(os.path.join(temp_folder, f))

            if final_xml_files:
                st.success(f"Tổng số hóa đơn: {len(final_xml_files)}")
                download_tar(temp_folder)


                status.update(label="Tải thành công !!!", expanded=True)
                return final_xml_files
            else:
                st.info("Không tìm thấy hóa đơn có thể tải !!!")
                status.update(label="Không tìm thấy hóa đơn có thể tải !!!", state="error")
                return None

        finally:
            if driver:
                driver.quit()
            
### TAB 2 FUNCTIONS
def extract_number_viettel(string):
    match = re.search(r'TAV(\d+)', string)
    if match:
        return match.group(1)
    else:
        return None

def download_icon_viettel(driver, action, wait, temp_folder):
    placeholder= st.empty()
    xml_files = []
    try:
        icons = wait.until(
            EC.presence_of_all_elements_located((By.XPATH, "//button[i[contains(@class, 'fa-info icon-info')]]"))
        )

        for icon in icons:
            action.move_to_element(icon).perform()
            driver.execute_script("arguments[0].click();", icon)
            time.sleep(3)
            
            invoice_form = driver.find_element(By.XPATH, "//div[@class='modal-content']")
            download_button = invoice_form.find_element(By.XPATH, "//button[span[text()='Tải xml']]")
            driver.execute_script("arguments[0].click();", download_button)
            time.sleep(3)

            downloaded_file = wait_for_download(temp_folder)
            if downloaded_file:
                shd = extract_number_viettel(os.path.basename(downloaded_file))
                placeholder.write_stream(stream_data(f"Đang tải hóa đơn số {shd} ..."))
                xml_file = shd + downloaded_file[downloaded_file.index('.xml'):]
                xml_files.append(xml_file)
                os.rename(os.path.join(temp_folder, downloaded_file), os.path.join(temp_folder, xml_file))

            close_button = invoice_form.find_element(By.XPATH, "//button[@class='close']")
            driver.execute_script("arguments[0].click();", close_button)
            time.sleep(2)
            
        if xml_files:
            return xml_files
        else:
            st.write_stream(stream_data("Không có hóa đơn nào được tải xuống"))
            return None

    except StaleElementReferenceException:
        return download_icon_viettel(driver, action, wait, temp_folder)
    except TimeoutException:
        st.write_stream(stream_data(f"Không tìm thấy hóa đơn"))
        return None

def handle_viettel_download(driver, action, wait, user, start_date, end_date, temp_folder):
    with st.status("Đang tải XML tự động ...", expanded=True) as status:
        try:
            driver.get('https://vinvoice.viettel.vn/account/login')

            wait.until(
                EC.presence_of_element_located((By.XPATH, '//form[@role="form"]'))
            )

            st.write_stream(stream_data((f"Đang đăng nhập tài khoản {user}...")))

            username = driver.find_element(By.ID, 'username')
            password = driver.find_element(By.NAME, 'password')

            username.send_keys(os.getenv('username_viettel'))
            password.send_keys(os.getenv('password_viettel'))
            password.send_keys(Keys.RETURN)

            qlhd = wait.until(
                EC.presence_of_element_located((By.XPATH,"//a[@href='/invoice-management/invoice']")))
            driver.execute_script("arguments[0].click();", qlhd)
            st.write_stream(stream_data(("Đang vào mục Quản Lý Hóa Đơn ...")))

            enter_dates(wait, start_date, end_date, btn_path="//input[@formcontrolname='datePicker']")

            search_btn = wait.until(
                EC.presence_of_element_located((By.XPATH, "//button[span[text()='Tìm kiếm']]"))
            )
            search_btn.click()
            st.write_stream(stream_data(("Đang tìm hóa đơn ...")))

            all_pages = driver.find_elements(By.XPATH, "//a[@class='page-link ng-star-inserted']")
            st.write_stream(stream_data((f"Tổng số trang: {len(all_pages)}")))
            
            final_xml_files = []
            page_index = 0
            while page_index < len(all_pages):
                try:
                    xml_files = download_icon_viettel(driver, action, wait, temp_folder)
                    if xml_files:
                        final_xml_files.extend(xml_files)
                    next_button = wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@aria-label='Next' and contains(@class, 'page-link')]"))
                    )
                    driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
                    driver.execute_script("arguments[0].click();", next_button)
                    page_index += 1  
                    time.sleep(2)  
                except TimeoutException:
                    break  
                except Exception as e:
                    st.write(f"Lỗi tải: {e}")
                    return None

            if final_xml_files:
                st.success(f"Tổng số hóa đơn: {len(final_xml_files)}")
                status.update(label="Tải thành công !!!", expanded=True)
                download_tar(temp_folder)
                return final_xml_files
            else:
                st.info("Không có hóa đơn được tìm thấy !!!")
                status.update(label="Không có hóa đơn được tìm thấy !!!", state="error")
                return None

        finally:
            if driver:
                driver.quit()  
    
### Streamlit State FUNCTIONS
def create():
    st.session_state['create_success'] = True

def download():
    st.session_state['download_success'] = True

def download_ptt():
    st.session_state['download_success_ptt'] = True
    
### MAIN FUNCTION
def main():
    tab1, tab2, tab3= st.tabs(['VNPT', 'Viettel','Phiếu Xuất Kho'])

    tab1.title("VNPT")
    tab2.title("Viettel")
    tab3.title("Phiếu Xuất Kho")

    with st.sidebar:
        xml_files = st.file_uploader("Nhập XML files", accept_multiple_files=True, type='xml')
        if len(xml_files) == 0:
            st.warning("Vui lòng tải tệp XML của bạn")
        else:
            st.success(f'Bạn đã tải thành công {len(xml_files)} tệp')
        st.divider()
            

    with tab1:   
        user = st.radio(
            "Hộ kinh doanh:",
            ["Trần Minh Đạt", "Nguyễn Thị Thanh Thúy"]
        )
        start_date, end_date = set_date(key1='vnpt_start', key2='vnpt_end')
        if st.button("Tải XML tự động", key="vnpt"):
            temp_folder = tempfile.mkdtemp()
            driver, action, wait = selenium_web_driver(temp_folder)
            handle_vnpt_download(driver, action, wait, user, start_date, end_date, temp_folder)

    with tab2:
        user = st.radio(
            "Hộ kinh doanh:",
            ["An Vinh"]
        )
        start_date, end_date= set_date(key1='viettel_start', key2='viettel_end')
        if st.button("Tải XML tự động", key="viettel"):    
            temp_folder = tempfile.mkdtemp()
            driver, action, wait = selenium_web_driver(temp_folder)
            handle_viettel_download(driver, action, wait, user, start_date, end_date, temp_folder)   

    with tab3:
        if xml_files:
            all_data = []
            for uploaded_file in xml_files:
                
                shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts, ggia, data= pxk_data_from_xml(uploaded_file)

                with st.container(border=True): 
                    df = display_pxk(shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts, ggia, data)
                    all_data.append((shdon, nmua, nmua_dc, nban, nban_dc, nban_mst, date, tbc, ts ,ggia, df))

            if 'create_success' not in st.session_state:
                st.session_state['create_success'] = False
            if 'download_success' not in st.session_state:
                st.session_state['download_success'] = False
            if 'download_success_ptt' not in st.session_state:
                st.session_state['download_success_ptt'] = False

            with st.sidebar:
                download_success_handler('download_success', 'Đang tải phiếu xuất kho ...')
                download_success_handler('download_success_ptt', 'Đang tải phiếu thu tiền ...')

                if not st.session_state.get('create_success', False):
                    if st.button('Tạo phiếu xuất kho và thu tiền', type='primary', key='btn', on_click=create):
                        pass
                else:
                    pxk_buffer, ptt_buffer = create_pxk(all_data)
                    st.success("Tạo phiếu xuất kho và thu tiền thành công")
                    
                    # PXK download button
                    create_download_button(
                        label="Tải phiếu xuất kho",
                        buffer=pxk_buffer,
                        file_name="PHIEU XUAT KHO.xlsx",
                        key="pxk"
                    )

                    # PTT download button
                    create_download_button(
                        label="Tải phiếu thu tiền",
                        buffer=ptt_buffer,
                        file_name="PHIEU THU TIEN.xlsx",
                        key="ptt"
                    )


if __name__ == "__main__":
    main()